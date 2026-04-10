from __future__ import annotations

import io
from datetime import date, datetime, timezone
from typing import Any, Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import extract, func
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models.catalog import Service, ServiceCatalog, VehicleModel
from app.models.quotes import Quote, QuoteLine, QuoteStatus
from app.models.suppliers import Supplier, SupplierPrice
from app.schemas.catalog import PaginatedResponse
from app.schemas.engine import CalculationInput, ScoringWeights, SupplierOption
from app.schemas.quotes import (
    QuoteCreate,
    QuoteListItem,
    QuoteRead,
    QuoteLineRead,
    QuoteSummary,
    QuoteStats,
    QuoteStatsByModel,
    QuoteUpdate,
)
from app.security import get_current_user, require_role
from app.services.pricing_engine import PricingEngine
from app.services.supplier_engine import SupplierEngine

router = APIRouter(prefix="/quotes", tags=["quotes"])


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _get_engine_config(db: Session) -> dict:
    """Read config_params from DB. Returns defaults if not present."""
    from app.models.config import ConfigParam

    defaults = {
        "technician_cost_hr": 156.25,
        "target_margin": 0.40,
        "scoring_weight_price": 0.50,
        "scoring_weight_time": 0.30,
        "scoring_weight_tc": 0.20,
    }
    params = db.query(ConfigParam).filter(ConfigParam.key.in_(defaults.keys())).all()
    result = dict(defaults)
    for p in params:
        result[p.key] = float(p.value)
    return result


def _get_quote_or_404(quote_id: str, db: Session) -> Quote:
    quote = (
        db.query(Quote)
        .options(joinedload(Quote.lines).joinedload(QuoteLine.service))
        .options(joinedload(Quote.model))
        .filter(Quote.id == quote_id, Quote.deleted_at.is_(None))
        .first()
    )
    if not quote:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cotización no encontrada",
        )
    return quote


def compute_quote_summary(lines: list[QuoteLine]) -> dict:
    valid_lines = [line for line in lines if line.margin_status != "no_data"]
    total_bjx = sum(line.total_bjx_cost for line in valid_lines)
    total_brame = sum(line.brame_price for line in valid_lines)
    total_margin_pesos = sum(line.margin_pesos for line in valid_lines)
    blended_pct = total_margin_pesos / total_brame if total_brame > 0 else 0.0
    return {
        "total_bjx_cost": total_bjx,
        "total_brame_price": total_brame,
        "blended_margin_pct": blended_pct,
        "blended_margin_pesos": total_margin_pesos,
        "ok_count": sum(1 for line in lines if line.margin_status == "ok"),
        "low_count": sum(1 for line in lines if line.margin_status == "low"),
        "critical_count": sum(1 for line in lines if line.margin_status == "critical"),
        "no_data_count": sum(1 for line in lines if line.margin_status == "no_data"),
    }


def _build_quote_read(quote: Quote) -> QuoteRead:
    """Build QuoteRead from a Quote ORM instance (with lines loaded)."""
    model_name = quote.model.name if quote.model else quote.model_id
    summary_data = compute_quote_summary(quote.lines)

    lines_read: list[QuoteLineRead] = []
    for line in quote.lines:
        service_name = line.service.name if line.service else line.service_id
        lines_read.append(
            QuoteLineRead(
                id=line.id,
                service_id=line.service_id,
                service_name=service_name,
                supplier_id=line.supplier_id,
                duration_hrs=line.duration_hrs,
                labor_cost=line.labor_cost,
                parts_cost=line.parts_cost,
                total_bjx_cost=line.total_bjx_cost,
                brame_price=line.brame_price,
                margin_pesos=line.margin_pesos,
                margin_pct=line.margin_pct,
                suggested_price=line.suggested_price,
                gap_vs_target=line.gap_vs_target,
                margin_status=line.margin_status,
                data_source=line.data_source,
            )
        )

    return QuoteRead(
        id=quote.id,
        quote_number=quote.quote_number,
        model_id=quote.model_id,
        model_name=model_name,
        created_by=quote.created_by,
        status=quote.status,
        technician_cost_hr=quote.technician_cost_hr,
        target_margin=quote.target_margin,
        notes=quote.notes,
        created_at=quote.created_at,
        lines=lines_read,
        summary=QuoteSummary(**summary_data),
    )


def _build_quote_list_item(quote: Quote) -> QuoteListItem:
    """Build QuoteListItem (without lines) from a Quote ORM instance."""
    model_name = quote.model.name if quote.model else quote.model_id
    summary_data = compute_quote_summary(quote.lines)
    return QuoteListItem(
        id=quote.id,
        quote_number=quote.quote_number,
        model_id=quote.model_id,
        model_name=model_name,
        created_by=quote.created_by,
        status=quote.status,
        technician_cost_hr=quote.technician_cost_hr,
        target_margin=quote.target_margin,
        notes=quote.notes,
        created_at=quote.created_at,
        summary=QuoteSummary(**summary_data),
    )


def _calculate_lines_for_quote(
    model_id: str,
    service_ids: list[str],
    technician_cost_hr: float,
    target_margin: float,
    db: Session,
) -> list[dict]:
    """
    Calculate pricing for each service_id using PricingEngine + SupplierEngine.
    Returns a list of dicts with QuoteLine field values.
    """
    cfg = _get_engine_config(db)
    weights = ScoringWeights(
        price_weight=cfg["scoring_weight_price"],
        time_weight=cfg["scoring_weight_time"],
        tc_weight=cfg["scoring_weight_tc"],
    )

    # Bulk-load catalog entries
    catalog_rows = (
        db.query(ServiceCatalog)
        .options(joinedload(ServiceCatalog.service))
        .filter(
            ServiceCatalog.model_id == model_id,
            ServiceCatalog.service_id.in_(service_ids),
            ServiceCatalog.is_current.is_(True),
        )
        .all()
    )
    catalog_by_service: dict[str, ServiceCatalog] = {c.service_id: c for c in catalog_rows}

    # Bulk-load supplier prices
    supplier_price_rows = (
        db.query(SupplierPrice)
        .options(joinedload(SupplierPrice.supplier))
        .filter(
            SupplierPrice.model_id == model_id,
            SupplierPrice.service_id.in_(service_ids),
            SupplierPrice.is_current.is_(True),
        )
        .all()
    )
    supplier_prices_by_service: dict[str, list[SupplierPrice]] = {}
    for sp in supplier_price_rows:
        supplier_prices_by_service.setdefault(sp.service_id, []).append(sp)

    pricing_engine = PricingEngine()
    supplier_engine = SupplierEngine()

    result_lines: list[dict] = []

    for sid in service_ids:
        catalog = catalog_by_service.get(sid)

        if catalog is None:
            # No catalog data → no_data line with zeros
            result_lines.append(
                {
                    "service_id": sid,
                    "supplier_id": None,
                    "duration_hrs": 0.0,
                    "labor_cost": 0.0,
                    "parts_cost": 0.0,
                    "total_bjx_cost": 0.0,
                    "brame_price": 0.0,
                    "margin_pesos": 0.0,
                    "margin_pct": 0.0,
                    "suggested_price": 0.0,
                    "gap_vs_target": 0.0,
                    "margin_status": "no_data",
                    "data_source": "no_data",
                }
            )
            continue

        # Build CalculationInput using best supplier prices (or 0 if none)
        sp_list = supplier_prices_by_service.get(sid, [])
        options: list[SupplierOption] = []
        for sp in sp_list:
            options.append(
                SupplierOption(
                    supplier_id=sp.supplier_id,
                    supplier_name=sp.supplier.name,
                    ref_cost=sp.ref_cost,
                    labor_cost=sp.labor_cost,
                    total_price=sp.total_price,
                    lead_time_days=sp.supplier.lead_time_days,
                    warranty_days=sp.supplier.warranty_days,
                )
            )

        scored = supplier_engine.score(options, weights)
        best_supplier = next((s for s in scored if s.recommended), None)

        brame_ref_actual = best_supplier.ref_cost if best_supplier else (catalog.bjx_parts_cost or 0.0)
        brame_total_actual = (
            best_supplier.total_price
            if best_supplier
            else (catalog.bjx_labor_cost or 0.0) + (catalog.bjx_parts_cost or 0.0)
        )

        inp = CalculationInput(
            model_id=model_id,
            service_id=sid,
            technician_cost_hr=technician_cost_hr,
            target_margin=target_margin,
            catalog_labor_cost=catalog.bjx_labor_cost,
            catalog_parts_cost=catalog.bjx_parts_cost,
            catalog_duration_hrs=catalog.duration_hrs,
            brame_ref_actual=brame_ref_actual,
            brame_total_actual=brame_total_actual,
        )

        calc = pricing_engine.calculate(inp)

        result_lines.append(
            {
                "service_id": sid,
                "supplier_id": best_supplier.supplier_id if best_supplier else None,
                "duration_hrs": calc.duration_hrs,
                "labor_cost": calc.labor_cost,
                "parts_cost": calc.parts_cost,
                "total_bjx_cost": calc.total_bjx_cost,
                "brame_price": calc.brame_price,
                "margin_pesos": calc.margin_pesos,
                "margin_pct": calc.margin_pct,
                "suggested_price": calc.suggested_price,
                "gap_vs_target": calc.gap_vs_target,
                "margin_status": calc.margin_status,
                "data_source": calc.data_source,
            }
        )

    return result_lines


# ---------------------------------------------------------------------------
# LIFECYCLE VALIDATION
# ---------------------------------------------------------------------------

# Valid transitions: {from_status: [allowed_to_statuses]}
_VALID_TRANSITIONS: dict[QuoteStatus, list[QuoteStatus]] = {
    QuoteStatus.draft: [QuoteStatus.confirmed, QuoteStatus.cancelled],
    QuoteStatus.confirmed: [QuoteStatus.invoiced, QuoteStatus.cancelled],
    QuoteStatus.invoiced: [QuoteStatus.cancelled],
    QuoteStatus.cancelled: [],
}

# Transitions that require admin role
_ADMIN_ONLY_TRANSITIONS: set[tuple[QuoteStatus, QuoteStatus]] = {
    (QuoteStatus.invoiced, QuoteStatus.cancelled),
}


def _validate_transition(
    current: QuoteStatus,
    new: QuoteStatus,
    user_role: str,
) -> None:
    allowed = _VALID_TRANSITIONS.get(current, [])
    if new not in allowed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Transición de estado inválida: {current.value} → {new.value}",
        )
    if (current, new) in _ADMIN_ONLY_TRANSITIONS and user_role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo un administrador puede realizar esta transición de estado",
        )


# ---------------------------------------------------------------------------
# POST /quotes
# ---------------------------------------------------------------------------


@router.post("", response_model=QuoteRead, status_code=status.HTTP_201_CREATED)
def create_quote(
    payload: QuoteCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_role(["admin", "operador"])),
):
    # 1. Read config defaults from DB
    cfg = _get_engine_config(db)
    technician_cost_hr = (
        payload.technician_cost_hr
        if payload.technician_cost_hr is not None
        else cfg["technician_cost_hr"]
    )
    target_margin = (
        payload.target_margin
        if payload.target_margin is not None
        else cfg["target_margin"]
    )

    # 2. Verify model_id exists
    vehicle_model = (
        db.query(VehicleModel)
        .filter(VehicleModel.id == payload.model_id, VehicleModel.deleted_at.is_(None))
        .first()
    )
    if not vehicle_model:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Modelo no encontrado: {payload.model_id}",
        )

    # 3. Generate quote_number: BJX-{YYYY}-{NNNN}
    current_year = datetime.now(timezone.utc).year
    year_count = (
        db.query(func.count(Quote.id))
        .filter(extract("year", Quote.created_at) == current_year)
        .scalar()
        or 0
    )
    sequence = year_count + 1
    quote_number = f"BJX-{current_year}-{str(sequence).zfill(4)}"

    # 4. Calculate lines using PricingEngine + SupplierEngine
    line_data_list = _calculate_lines_for_quote(
        model_id=payload.model_id,
        service_ids=payload.service_ids,
        technician_cost_hr=technician_cost_hr,
        target_margin=target_margin,
        db=db,
    )

    # 5. Create Quote
    quote = Quote(
        quote_number=quote_number,
        model_id=payload.model_id,
        created_by=current_user.email,
        status=QuoteStatus.draft,
        technician_cost_hr=technician_cost_hr,
        target_margin=target_margin,
        notes=payload.notes,
    )
    db.add(quote)
    db.flush()  # get quote.id before creating lines

    # 6. Create QuoteLine records
    for line_data in line_data_list:
        line = QuoteLine(
            quote_id=quote.id,
            **line_data,
        )
        db.add(line)

    db.commit()

    # 7. Reload with relationships
    db.refresh(quote)
    quote = _get_quote_or_404(quote.id, db)
    return _build_quote_read(quote)


# ---------------------------------------------------------------------------
# GET /quotes/stats  — must be before GET /quotes/{id}
# ---------------------------------------------------------------------------


@router.get("/stats", response_model=QuoteStats)
def get_quote_stats(
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    # Only confirmed + invoiced for margin metrics
    metric_statuses = [QuoteStatus.confirmed, QuoteStatus.invoiced]

    base_q = db.query(Quote).filter(Quote.deleted_at.is_(None))
    metric_q = base_q.filter(Quote.status.in_(metric_statuses))

    if from_date:
        base_q = base_q.filter(Quote.created_at >= datetime(from_date.year, from_date.month, from_date.day, tzinfo=timezone.utc))
        metric_q = metric_q.filter(Quote.created_at >= datetime(from_date.year, from_date.month, from_date.day, tzinfo=timezone.utc))
    if to_date:
        to_dt = datetime(to_date.year, to_date.month, to_date.day, 23, 59, 59, tzinfo=timezone.utc)
        base_q = base_q.filter(Quote.created_at <= to_dt)
        metric_q = metric_q.filter(Quote.created_at <= to_dt)

    total_quotes = base_q.count()

    # by_status count
    status_counts = (
        db.query(Quote.status, func.count(Quote.id))
        .filter(Quote.deleted_at.is_(None))
        .group_by(Quote.status)
        .all()
    )
    by_status: dict[str, int] = {s.value: 0 for s in QuoteStatus}
    for s, cnt in status_counts:
        by_status[s.value] = cnt

    # Weighted blended margin across confirmed + invoiced quotes
    # We need sum of margin_pesos and sum of brame_price across their lines
    from app.models.quotes import QuoteLine as QL

    agg = (
        db.query(
            func.sum(QL.margin_pesos).label("total_margin_pesos"),
            func.sum(QL.brame_price).label("total_brame_price"),
        )
        .join(Quote, Quote.id == QL.quote_id)
        .filter(
            Quote.deleted_at.is_(None),
            Quote.status.in_(metric_statuses),
            QL.margin_status != "no_data",
        )
    )
    if from_date:
        agg = agg.filter(Quote.created_at >= datetime(from_date.year, from_date.month, from_date.day, tzinfo=timezone.utc))
    if to_date:
        agg = agg.filter(Quote.created_at <= datetime(to_date.year, to_date.month, to_date.day, 23, 59, 59, tzinfo=timezone.utc))

    agg_result = agg.first()
    total_margin_pesos = float(agg_result.total_margin_pesos or 0.0)
    total_brame_price = float(agg_result.total_brame_price or 0.0)
    avg_blended_margin_pct = (
        total_margin_pesos / total_brame_price if total_brame_price > 0 else 0.0
    )

    # critical and ok counts (quotes where blended margin is critical or ok)
    # Use per-quote summary; simplify by counting quotes that have ANY critical / all ok lines
    metric_quotes = metric_q.options(
        joinedload(Quote.lines)
    ).all()

    critical_quotes_count = 0
    ok_quotes_count = 0
    for q in metric_quotes:
        summary = compute_quote_summary(q.lines)
        if summary["critical_count"] > 0:
            critical_quotes_count += 1
        elif summary["ok_count"] == len([l for l in q.lines if l.margin_status != "no_data"]) and summary["ok_count"] > 0:
            ok_quotes_count += 1

    # by_model — count of confirmed+invoiced quotes per model
    model_counts = (
        db.query(Quote.model_id, func.count(Quote.id).label("cnt"))
        .filter(
            Quote.deleted_at.is_(None),
            Quote.status.in_(metric_statuses),
        )
        .group_by(Quote.model_id)
        .order_by(func.count(Quote.id).desc())
        .all()
    )

    model_ids = [row[0] for row in model_counts]
    models_map: dict[str, str] = {}
    if model_ids:
        model_objs = db.query(VehicleModel).filter(VehicleModel.id.in_(model_ids)).all()
        models_map = {m.id: m.name for m in model_objs}

    by_model = [
        QuoteStatsByModel(
            model_id=mid,
            model_name=models_map.get(mid, mid),
            count=cnt,
        )
        for mid, cnt in model_counts
    ]

    return QuoteStats(
        period={
            "from": from_date.isoformat() if from_date else None,
            "to": to_date.isoformat() if to_date else None,
        },
        total_quotes=total_quotes,
        by_status=by_status,
        avg_blended_margin_pct=avg_blended_margin_pct,
        critical_quotes_count=critical_quotes_count,
        ok_quotes_count=ok_quotes_count,
        by_model=by_model,
    )


# ---------------------------------------------------------------------------
# GET /quotes
# ---------------------------------------------------------------------------


@router.get("", response_model=PaginatedResponse[QuoteListItem])
def list_quotes(
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=100),
    status_filter: Optional[QuoteStatus] = Query(default=None, alias="status"),
    model_id: Optional[str] = Query(default=None),
    from_date: Optional[date] = Query(default=None, alias="from"),
    to_date: Optional[date] = Query(default=None, alias="to"),
    sort: Literal["created_at", "margin_pct", "quote_number"] = Query(default="created_at"),
    order: Literal["asc", "desc"] = Query(default="desc"),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    from sqlalchemy import case, literal

    query = (
        db.query(Quote)
        .options(joinedload(Quote.lines), joinedload(Quote.model))
        .filter(Quote.deleted_at.is_(None))
    )

    if status_filter is not None:
        query = query.filter(Quote.status == status_filter)
    if model_id:
        query = query.filter(Quote.model_id == model_id)
    if from_date:
        query = query.filter(
            Quote.created_at >= datetime(from_date.year, from_date.month, from_date.day, tzinfo=timezone.utc)
        )
    if to_date:
        query = query.filter(
            Quote.created_at <= datetime(to_date.year, to_date.month, to_date.day, 23, 59, 59, tzinfo=timezone.utc)
        )

    # Ordering
    if sort == "margin_pct":
        # Subquery: blended margin per quote = sum(margin_pesos) / sum(brame_price)
        subq = (
            db.query(
                QuoteLine.quote_id,
                (
                    func.sum(QuoteLine.margin_pesos) / func.nullif(func.sum(QuoteLine.brame_price), 0)
                ).label("blended_margin"),
            )
            .filter(QuoteLine.margin_status != "no_data")
            .group_by(QuoteLine.quote_id)
            .subquery()
        )
        query = query.outerjoin(subq, subq.c.quote_id == Quote.id)
        sort_expr = subq.c.blended_margin
    elif sort == "quote_number":
        sort_expr = Quote.quote_number
    else:
        sort_expr = Quote.created_at

    if order == "desc":
        query = query.order_by(sort_expr.desc())
    else:
        query = query.order_by(sort_expr.asc())

    total = query.count()
    quotes = query.offset((page - 1) * size).limit(size).all()

    items = [_build_quote_list_item(q) for q in quotes]

    return PaginatedResponse[QuoteListItem](
        items=items,
        total=total,
        page=page,
        size=size,
    )


# ---------------------------------------------------------------------------
# GET /quotes/{id}
# ---------------------------------------------------------------------------


@router.get("/{quote_id}", response_model=QuoteRead)
def get_quote(
    quote_id: str,
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    quote = _get_quote_or_404(quote_id, db)
    return _build_quote_read(quote)


# ---------------------------------------------------------------------------
# PUT /quotes/{id}
# ---------------------------------------------------------------------------


@router.put("/{quote_id}", response_model=QuoteRead)
def update_quote(
    quote_id: str,
    payload: QuoteUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(require_role(["admin", "operador"])),
):
    quote = _get_quote_or_404(quote_id, db)

    if payload.status is not None:
        _validate_transition(quote.status, payload.status, current_user.role.value)
        quote.status = payload.status

    if payload.notes is not None:
        quote.notes = payload.notes

    db.commit()
    db.refresh(quote)
    quote = _get_quote_or_404(quote.id, db)
    return _build_quote_read(quote)


# ---------------------------------------------------------------------------
# DELETE /quotes/{id}  — soft delete
# ---------------------------------------------------------------------------


@router.delete("/{quote_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_quote(
    quote_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_role(["admin", "operador"])),
):
    quote = _get_quote_or_404(quote_id, db)

    if quote.status == QuoteStatus.invoiced:
        if current_user.role.value != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Solo un administrador puede cancelar una cotización facturada",
            )

    if quote.status not in (QuoteStatus.draft, QuoteStatus.confirmed, QuoteStatus.invoiced):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La cotización ya está cancelada",
        )

    quote.status = QuoteStatus.cancelled
    quote.deleted_at = datetime.now(timezone.utc)
    db.commit()


# ---------------------------------------------------------------------------
# GET /quotes/{id}/export
# ---------------------------------------------------------------------------


@router.get("/{quote_id}/export")
def export_quote(
    quote_id: str,
    format: Literal["pdf", "xlsx"] = Query(...),
    db: Session = Depends(get_db),
    _user=Depends(get_current_user),
):
    quote = _get_quote_or_404(quote_id, db)

    if quote.status not in (QuoteStatus.confirmed, QuoteStatus.invoiced):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden exportar cotizaciones confirmadas o facturadas",
        )

    today_str = date.today().isoformat()

    if format == "pdf":
        return _export_pdf(quote, today_str)
    else:
        return _export_xlsx(quote, today_str)


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

_STATUS_COLORS_HEX = {
    "ok": "#10B981",
    "low": "#F97316",
    "critical": "#EF4444",
    "no_data": "#9CA3AF",
}


def _export_pdf(quote: Quote, today_str: str) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=0.5 * inch, rightMargin=0.5 * inch)
    styles = getSampleStyleSheet()

    model_name = quote.model.name if quote.model else quote.model_id
    summary = compute_quote_summary(quote.lines)

    elements = []

    # Header
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=4,
    )
    elements.append(Paragraph(f"Cotización {quote.quote_number}", title_style))
    elements.append(Paragraph(f"Modelo: {model_name}", styles["Normal"]))
    elements.append(Paragraph(f"Fecha: {today_str}", styles["Normal"]))
    elements.append(Paragraph(f"Estado: {quote.status.value}", styles["Normal"]))
    elements.append(Spacer(1, 0.2 * inch))

    # Table header
    table_headers = [
        "Concepto",
        "Duración (hrs)",
        "Costo MO",
        "Costo Ref.",
        "Costo Total BJX",
        "Precio Brame",
        "Margen $",
        "Margen %",
    ]
    table_data = [table_headers]

    # Table rows
    row_colors: list[tuple[int, colors.Color]] = []
    for i, line in enumerate(quote.lines, start=1):
        svc_name = line.service.name if line.service else line.service_id
        row = [
            svc_name,
            f"{line.duration_hrs:.2f}",
            f"${line.labor_cost:,.2f}",
            f"${line.parts_cost:,.2f}",
            f"${line.total_bjx_cost:,.2f}",
            f"${line.brame_price:,.2f}",
            f"${line.margin_pesos:,.2f}",
            f"{line.margin_pct * 100:.1f}%",
        ]
        table_data.append(row)

        # Map margin_status to reportlab color
        status_color_map = {
            "ok": colors.HexColor("#10B981"),
            "low": colors.HexColor("#F97316"),
            "critical": colors.HexColor("#EF4444"),
            "no_data": colors.HexColor("#9CA3AF"),
        }
        row_colors.append((i, status_color_map.get(line.margin_status, colors.white)))

    table = Table(
        table_data,
        colWidths=[2.0 * inch, 0.9 * inch, 0.9 * inch, 0.9 * inch, 1.1 * inch, 1.0 * inch, 0.9 * inch, 0.8 * inch],
        repeatRows=1,
    )

    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    # Add semaphore background on margin% column (col 7)
    for row_idx, col_color in row_colors:
        base_style.append(("BACKGROUND", (7, row_idx), (7, row_idx), col_color))
        base_style.append(("TEXTCOLOR", (7, row_idx), (7, row_idx), colors.white))

    table.setStyle(TableStyle(base_style))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))

    # Summary section
    elements.append(Paragraph("Resumen", styles["Heading2"]))
    summary_data_table = [
        ["Costo Total BJX", f"${summary['total_bjx_cost']:,.2f}"],
        ["Precio Total Brame", f"${summary['total_brame_price']:,.2f}"],
        ["Margen Total $", f"${summary['blended_margin_pesos']:,.2f}"],
        ["Margen Blended %", f"{summary['blended_margin_pct'] * 100:.1f}%"],
        ["Líneas OK / Bajo / Crítico / Sin datos",
         f"{summary['ok_count']} / {summary['low_count']} / {summary['critical_count']} / {summary['no_data_count']}"],
    ]
    summary_table = Table(summary_data_table, colWidths=[3 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.2 * inch))

    # Footer / parameters
    elements.append(Paragraph(
        f"Parámetros: Costo técnico/hr = ${quote.technician_cost_hr:.2f} | "
        f"Margen objetivo = {quote.target_margin * 100:.1f}%",
        styles["Normal"],
    ))
    elements.append(Paragraph(f"Generado: {today_str} | Creado por: {quote.created_by}", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    filename = f"BJX-{quote.quote_number}-{today_str}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_xlsx(quote: Quote, today_str: str) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = quote.quote_number

    model_name = quote.model.name if quote.model else quote.model_id
    summary = compute_quote_summary(quote.lines)

    # Header info
    ws["A1"] = f"Cotización: {quote.quote_number}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Modelo: {model_name}"
    ws["A3"] = f"Fecha: {today_str}"
    ws["A4"] = f"Estado: {quote.status.value}"
    ws["A5"] = f"Creado por: {quote.created_by}"

    header_row = 7
    headers = [
        "Concepto",
        "Duración (hrs)",
        "Costo MO",
        "Costo Ref.",
        "Costo Total BJX",
        "Precio Brame",
        "Margen $",
        "Margen %",
        "Estado Margen",
        "Fuente",
    ]

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Semaphore colors (ARGB)
    status_fill_map = {
        "ok": PatternFill(start_color="FF10B981", end_color="FF10B981", fill_type="solid"),
        "low": PatternFill(start_color="FFF97316", end_color="FFF97316", fill_type="solid"),
        "critical": PatternFill(start_color="FFEF4444", end_color="FFEF4444", fill_type="solid"),
        "no_data": PatternFill(start_color="FF9CA3AF", end_color="FF9CA3AF", fill_type="solid"),
    }
    white_font = Font(color="FFFFFF", bold=True)

    for line_idx, line in enumerate(quote.lines, start=1):
        row = header_row + line_idx
        svc_name = line.service.name if line.service else line.service_id

        values = [
            svc_name,
            line.duration_hrs,
            line.labor_cost,
            line.parts_cost,
            line.total_bjx_cost,
            line.brame_price,
            line.margin_pesos,
            line.margin_pct,
            line.margin_status,
            line.data_source,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Format monetary columns
            if col_idx in (3, 4, 5, 6, 7):
                cell.number_format = '"$"#,##0.00'
            # Format pct column
            if col_idx == 8:
                cell.number_format = "0.0%"
                # Apply semaphore color
                fill = status_fill_map.get(line.margin_status)
                if fill:
                    cell.fill = fill
                    cell.font = white_font

        # Alternate row background for non-colored cells
        if line_idx % 2 == 0:
            alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            for col_idx in range(1, 11):
                if col_idx != 8:  # skip semaphore column
                    ws.cell(row=row, column=col_idx).fill = alt_fill

    # Summary section
    summary_start_row = header_row + len(quote.lines) + 2
    ws.cell(row=summary_start_row, column=1, value="RESUMEN").font = Font(bold=True, size=12)

    summary_rows = [
        ("Costo Total BJX", summary["total_bjx_cost"]),
        ("Precio Total Brame", summary["total_brame_price"]),
        ("Margen Total $", summary["blended_margin_pesos"]),
        ("Margen Blended %", summary["blended_margin_pct"]),
        (f"OK / Bajo / Crítico / Sin datos",
         f"{summary['ok_count']} / {summary['low_count']} / {summary['critical_count']} / {summary['no_data_count']}"),
    ]

    bold_font = Font(bold=True)
    for i, (label, value) in enumerate(summary_rows):
        r = summary_start_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = bold_font
        cell = ws.cell(row=r, column=2, value=value)
        if isinstance(value, float) and label.endswith("%"):
            cell.number_format = "0.0%"
        elif isinstance(value, float):
            cell.number_format = '"$"#,##0.00'

    # Parameters footer
    params_row = summary_start_row + len(summary_rows) + 2
    ws.cell(
        row=params_row,
        column=1,
        value=(
            f"Parámetros: Costo técnico/hr = ${quote.technician_cost_hr:.2f} | "
            f"Margen objetivo = {quote.target_margin * 100:.1f}%"
        ),
    )
    ws.cell(row=params_row + 1, column=1, value=f"Generado: {today_str}")

    # Column widths
    col_widths = [35, 14, 12, 12, 15, 14, 12, 12, 15, 12]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"BJX-{quote.quote_number}-{today_str}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
