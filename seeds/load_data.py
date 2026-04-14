"""
seeds/load_data.py
==================
Script de seed idempotente para BJX-Atlas-Api.

Carga datos desde:
  - BJX_Calculadora_Brame_v1.xlsx  → hoja Datos_Comparativo
    → service_catalog (costos BJX reales de columnas BJX_MO_CATÁLOGO / BJX_REF_CATÁLOGO)
    → supplier_prices (precios BRAME actuales e históricos)

Uso:
    python seeds/load_data.py
    python seeds/load_data.py --dry-run
    python seeds/load_data.py --reset      # borra SQLite y recrea schema (solo dev)
    python seeds/load_data.py --reset --dry-run
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import re
import math
from datetime import date
from typing import Optional

import openpyxl

from app.database import SessionLocal, engine, Base
from app.models.catalog import VehicleModel, Service, ServiceCatalog
from app.models.suppliers import Supplier, SupplierPrice
from app.models.config import ConfigParam
from app.models.users import User, Role

# ---------------------------------------------------------------------------
# Columnas de la hoja Datos_Comparativo (0-based)
# ---------------------------------------------------------------------------

TARGET_SHEET     = "Datos_Comparativo"

COL_MODELO       = 0
COL_SERVICIO     = 1
COL_REF_PASADO   = 2   # BRAME_REF_PASADO ($)
COL_MO_PASADO    = 3   # BRAME_MO_PASADO ($)
COL_TOTAL_PASADO = 4   # BRAME_TOTAL_PASADO ($)
COL_REF_ACTUAL   = 5   # BRAME_REF_ACTUAL ($)
COL_MO_ACTUAL    = 6   # BRAME_MO_ACTUAL ($)
COL_TOTAL_ACTUAL = 7   # BRAME_TOTAL_ACTUAL ($)
# cols 8-10 = margen% / costo_est / margen$ → ignorados (los calcula el motor)
COL_BJX_MO       = 11  # BJX_MO_CATÁLOGO ($)
COL_BJX_REF      = 12  # BJX_REF_CATÁLOGO ($)
COL_DURACION     = 13  # BJX_DURACION_HRS

# ---------------------------------------------------------------------------
# Categorías de servicio
# ---------------------------------------------------------------------------

SERVICE_CATEGORY_MAP: dict[str, str] = {
    "AFINACIÓN":                                    "motor",
    "AFINACION":                                    "motor",
    "AJUSTE DE BALATAS":                            "frenos",
    "ALINEACIÓN":                                   "suspension",
    "ALINEACION":                                   "suspension",
    "CAMBIO DE AMORTIGUADORES DELANTEROS":          "suspension",
    "CAMBIO DE ANTICONGELANTE":                     "motor",
    "CAMBIO DE BALATAS DELANTERAS":                 "frenos",
    "CAMBIO DE BALATAS TRASERAS":                   "frenos",
    "CAMBIO DE BASES DE AMORTIGUADORES DELANTEROS": "suspension",
    "CAMBIO DE BATERÍA":                            "electrico",
    "CAMBIO DE BATERIA":                            "electrico",
    "CAMBIO DE DEPÓSITO DE ANTICONGELANTE":         "motor",
    "CAMBIO DE DEPOSITO DE ANTICONGELANTE":         "motor",
    "CAMBIO DE HORQUILLAS DELANTERAS":              "suspension",
}

# ---------------------------------------------------------------------------
# Normalización de nombres de modelo
# ---------------------------------------------------------------------------

NORMALIZATIONS: dict[str, str] = {
    "NISSAN - MARCH":            "NISSAN MARCH",
    "CHEVROLET - TORNADO PICK UP": "CHEVROLET - TORNADO PU",
    "VOLKSWAGEN - SAVEIRO ":     "VOLKSWAGEN - SAVEIRO",
}

# ---------------------------------------------------------------------------
# Config params por defecto
# ---------------------------------------------------------------------------

DEFAULTS: dict[str, tuple[str, str]] = {
    "technician_cost_hr":    ("156.25", "Costo por hora del técnico BJX en MXN"),
    "target_margin":         ("0.40",   "Margen objetivo por defecto (40%)"),
    "iva_rate":              ("0.16",   "Tasa IVA"),
    "overhead_rate":         ("0.15",   "Tasa de overhead operativo"),
    "scoring_weight_price":  ("0.50",   "Peso del precio en scoring de proveedores"),
    "scoring_weight_time":   ("0.30",   "Peso del tiempo de entrega en scoring"),
    "scoring_weight_tc":     ("0.20",   "Peso de términos y condiciones en scoring"),
}

# ---------------------------------------------------------------------------
# Helpers de parsing
# ---------------------------------------------------------------------------

def parse_currency(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        return float(value)
    text = re.sub(r"[MX$\s,]", "", str(value).strip())
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_model_name(name: str) -> str:
    stripped = name.strip()
    if stripped in NORMALIZATIONS:
        return NORMALIZATIONS[stripped]
    for key, val in NORMALIZATIONS.items():
        if key.strip() == stripped:
            return val
    return stripped


def infer_service_category(name: str) -> str:
    key = name.upper().strip()
    # Exact match first
    if key in SERVICE_CATEGORY_MAP:
        return SERVICE_CATEGORY_MAP[key]
    # Keyword fallback
    if "BALATA" in key or "FRENO" in key:
        return "frenos"
    if "AMORTIGUADOR" in key or "ALINEAC" in key or "HORQUILLA" in key or "SUSPENSION" in key:
        return "suspension"
    if "ELECTR" in key or "BATER" in key or "FOCO" in key:
        return "electrico"
    if "NEUMATICO" in key or "LLANTA" in key or "RUEDA" in key:
        return "neumaticos"
    if "MOTOR" in key or "AFIN" in key or "ACEITE" in key or "ANTICONGELANTE" in key:
        return "motor"
    return "otros"

# ---------------------------------------------------------------------------
# Upsert helpers
# ---------------------------------------------------------------------------

def get_or_create_model(db, name: str) -> VehicleModel:
    normalized = normalize_model_name(name)
    key = normalized.lower().strip()
    for m in db.query(VehicleModel).all():
        if m.name.lower().strip() == key:
            return m
    brand = None
    if " - " in normalized:
        brand = normalized.split(" - ")[0].strip()
    elif " " in normalized:
        brand = normalized.split(" ")[0].strip()
    model = VehicleModel(name=normalized, brand=brand, active=True)
    db.add(model)
    db.flush()
    return model


def get_or_create_service(db, name: str) -> Service:
    key = name.lower().strip()
    for s in db.query(Service).all():
        if s.name.lower().strip() == key:
            return s
    category = infer_service_category(name)
    service = Service(name=name.strip(), category=category, active=True)
    db.add(service)
    db.flush()
    return service

# ---------------------------------------------------------------------------
# Seed: config params
# ---------------------------------------------------------------------------

def seed_config_params(db, dry_run: bool = False) -> int:
    inserted = 0
    for key, (value, description) in DEFAULTS.items():
        if db.query(ConfigParam).filter(ConfigParam.key == key).first():
            continue
        if dry_run:
            print(f"  [DRY-RUN] config: {key} = {value}")
        else:
            db.add(ConfigParam(key=key, value=value, description=description))
            inserted += 1
    if not dry_run and inserted:
        db.flush()
    return inserted

# ---------------------------------------------------------------------------
# Seed: Excel Brame — hoja Datos_Comparativo
# ---------------------------------------------------------------------------

def load_brame_excel(db, filepath: str, dry_run: bool = False) -> dict:
    stats = {
        "models": 0, "services": 0,
        "catalog_entries": 0,
        "brame_prices_current": 0, "brame_prices_past": 0,
        "errors": 0,
    }

    if not os.path.exists(filepath):
        print(f"[WARN] Archivo no encontrado: {filepath}")
        return stats

    print(f"\n[INFO] Leyendo: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    print(f"[INFO] Hojas disponibles: {wb.sheetnames}")

    # Buscar hoja cuyo nombre contenga TARGET_SHEET (ignora emojis de prefijo)
    matched_sheet = next(
        (s for s in wb.sheetnames if TARGET_SHEET in s),
        None
    )
    if not matched_sheet:
        print(f"[ERROR] Hoja con '{TARGET_SHEET}' no encontrada. Hojas: {wb.sheetnames}")
        stats["errors"] += 1
        return stats

    ws = wb[matched_sheet]
    print(f"[INFO] Procesando hoja '{matched_sheet}'")

    # Leer encabezado (fila 1) para validar columnas
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"[INFO] Encabezado ({len(header)} cols): {header}")

    # Asegurar proveedor BRAME
    brame = None
    if not dry_run:
        brame = db.query(Supplier).filter(Supplier.name.ilike("brame")).first()
        if not brame:
            brame = Supplier(name="BRAME", lead_time_days=1, warranty_days=30, active=True)
            db.add(brame)
            db.flush()
            print(f"  [OK] Proveedor BRAME creado")
        else:
            print(f"  [INFO] Proveedor BRAME ya existe")

    today = date.today()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Ignorar filas vacías
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        try:
            # Extraer valores por índice explícito
            raw_modelo   = str(row[COL_MODELO]   or "").strip()
            raw_servicio = str(row[COL_SERVICIO]  or "").strip()

            if not raw_modelo or not raw_servicio:
                continue
            # Saltar filas que sean encabezados duplicados
            if raw_modelo.lower() in ("modelo", "vehículo", "vehiculo"):
                continue

            bjx_mo    = parse_currency(row[COL_BJX_MO]   if len(row) > COL_BJX_MO   else None)
            bjx_ref   = parse_currency(row[COL_BJX_REF]  if len(row) > COL_BJX_REF  else None)
            duration  = parse_float(row[COL_DURACION]     if len(row) > COL_DURACION  else None) or 0.0

            brame_mo_actual    = parse_currency(row[COL_MO_ACTUAL]    if len(row) > COL_MO_ACTUAL    else None)
            brame_ref_actual   = parse_currency(row[COL_REF_ACTUAL]   if len(row) > COL_REF_ACTUAL   else None)
            brame_total_actual = parse_currency(row[COL_TOTAL_ACTUAL] if len(row) > COL_TOTAL_ACTUAL else None)

            brame_mo_pasado    = parse_currency(row[COL_MO_PASADO]    if len(row) > COL_MO_PASADO    else None)
            brame_ref_pasado   = parse_currency(row[COL_REF_PASADO]   if len(row) > COL_REF_PASADO   else None)
            brame_total_pasado = parse_currency(row[COL_TOTAL_PASADO] if len(row) > COL_TOTAL_PASADO else None)

            if dry_run:
                category = infer_service_category(raw_servicio)
                print(
                    f"  [DRY-RUN] row {row_idx}: "
                    f"modelo='{normalize_model_name(raw_modelo)}' "
                    f"servicio='{raw_servicio}' cat={category} "
                    f"bjx_mo={bjx_mo} bjx_ref={bjx_ref} dur={duration} "
                    f"brame_total_actual={brame_total_actual} brame_total_pasado={brame_total_pasado}"
                )
                stats["catalog_entries"] += 1
                if brame_total_actual is not None:
                    stats["brame_prices_current"] += 1
                if brame_total_pasado is not None:
                    stats["brame_prices_past"] += 1
                continue

            # Upsert modelo y servicio
            vehicle_model = get_or_create_model(db, raw_modelo)
            service       = get_or_create_service(db, raw_servicio)

            # ServiceCatalog — costos BJX reales (NULLs válidos → "estimated")
            existing_catalog = db.query(ServiceCatalog).filter(
                ServiceCatalog.model_id  == vehicle_model.id,
                ServiceCatalog.service_id == service.id,
                ServiceCatalog.is_current.is_(True),
            ).first()
            if not existing_catalog:
                db.add(ServiceCatalog(
                    model_id=vehicle_model.id,
                    service_id=service.id,
                    bjx_labor_cost=bjx_mo,    # columna BJX_MO_CATÁLOGO
                    bjx_parts_cost=bjx_ref,   # columna BJX_REF_CATÁLOGO
                    duration_hrs=duration,
                    source="brame_xlsx",
                    is_current=True,
                    updated_by="seed",
                ))
                stats["catalog_entries"] += 1

            # SupplierPrice ACTUAL (is_current=True)
            if brame_total_actual is not None:
                existing_cur = db.query(SupplierPrice).filter(
                    SupplierPrice.supplier_id == brame.id,
                    SupplierPrice.service_id  == service.id,
                    SupplierPrice.model_id    == vehicle_model.id,
                    SupplierPrice.is_current.is_(True),
                ).first()
                if not existing_cur:
                    db.add(SupplierPrice(
                        supplier_id=brame.id,
                        service_id=service.id,
                        model_id=vehicle_model.id,
                        ref_cost=brame_ref_actual   or 0.0,
                        labor_cost=brame_mo_actual  or 0.0,
                        total_price=brame_total_actual,
                        price_date=today,
                        is_current=True,
                    ))
                    stats["brame_prices_current"] += 1

            # SupplierPrice PASADO (is_current=False)
            if brame_total_pasado is not None:
                existing_past = db.query(SupplierPrice).filter(
                    SupplierPrice.supplier_id == brame.id,
                    SupplierPrice.service_id  == service.id,
                    SupplierPrice.model_id    == vehicle_model.id,
                    SupplierPrice.is_current.is_(False),
                ).first()
                if not existing_past:
                    db.add(SupplierPrice(
                        supplier_id=brame.id,
                        service_id=service.id,
                        model_id=vehicle_model.id,
                        ref_cost=brame_ref_pasado   or 0.0,
                        labor_cost=brame_mo_pasado  or 0.0,
                        total_price=brame_total_pasado,
                        price_date=today,
                        is_current=False,
                    ))
                    stats["brame_prices_past"] += 1

        except Exception as exc:
            print(f"  [ERROR] row {row_idx}: {exc}")
            stats["errors"] += 1
            continue

    if not dry_run:
        try:
            db.flush()
        except Exception as exc:
            print(f"[ERROR] flush: {exc}")
            db.rollback()
            stats["errors"] += 1

    if not dry_run:
        stats["models"]   = db.query(VehicleModel).count()
        stats["services"] = db.query(Service).count()

    return stats

# ---------------------------------------------------------------------------
# Seed: usuario admin
# ---------------------------------------------------------------------------

def seed_admin_user(db, dry_run: bool = False) -> bool:
    from app.security import hash_password
    email    = os.getenv("ADMIN_EMAIL",    "admin@bjx.com")
    password = os.getenv("ADMIN_PASSWORD", "Admin1234")
    if db.query(User).filter(User.email == email).first():
        print(f"  [INFO] Usuario admin ya existe: {email}")
        return False
    if dry_run:
        print(f"  [DRY-RUN] Crearía admin: {email}")
        return False
    db.add(User(
        email=email,
        hashed_password=hash_password(password),
        role=Role.admin,
        active=True,
    ))
    db.flush()
    print(f"  [OK] Usuario admin creado: {email} / {password}")
    return True

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Seed de datos para BJX-Atlas-Api")
    parser.add_argument("--dry-run", action="store_true", help="No escribe en BD")
    parser.add_argument("--reset",   action="store_true", help="Borra y recrea la BD (solo SQLite/dev)")
    args = parser.parse_args()
    dry_run: bool = args.dry_run

    if dry_run:
        print("=" * 60)
        print("MODO DRY-RUN: no se escribirá nada en la base de datos")
        print("=" * 60)

    # --reset: eliminar SQLite y recrear schema
    if args.reset and not dry_run:
        db_url = os.getenv("DATABASE_URL", "")
        if "sqlite" in db_url:
            # Extraer ruta del archivo
            db_path = db_url.split("sqlite:///")[-1].lstrip("./")
            # Manejar rutas relativas
            if not os.path.isabs(db_path):
                project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                db_path = os.path.join(project_root, db_path)
            if os.path.exists(db_path):
                os.remove(db_path)
                print(f"[RESET] Eliminado: {db_path}")
            else:
                print(f"[RESET] No existía: {db_path}")
        print("[RESET] Recreando schema...")
        Base.metadata.drop_all(bind=engine)
        Base.metadata.create_all(bind=engine)
        print("[RESET] Schema listo.")
    else:
        # Solo crear tablas si no existen (normal dev)
        Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    total_errors = 0

    try:
        # 1. Config params
        print("\n[STEP 1] Config params...")
        n_cfg = seed_config_params(db, dry_run=dry_run)
        if not dry_run:
            db.commit()
        print(f"  => {n_cfg} parámetros insertados")

        # 2. Excel Brame
        brame_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "context", "BJX_Calculadora_Brame_v1.xlsx",
        )
        print("\n[STEP 2] Carga Excel Brame...")
        stats = load_brame_excel(db, brame_path, dry_run=dry_run)
        if not dry_run:
            try:
                db.commit()
            except Exception as exc:
                print(f"[ERROR] Commit Excel: {exc}")
                db.rollback()
                total_errors += 1
        total_errors += stats.get("errors", 0)

        # 3. Usuario admin
        print("\n[STEP 3] Usuario admin...")
        seed_admin_user(db, dry_run=dry_run)
        if not dry_run:
            try:
                db.commit()
            except Exception as exc:
                print(f"[ERROR] Commit admin: {exc}")
                db.rollback()
                total_errors += 1

        # 4. Resumen
        print("\n" + "=" * 60)
        print("RESUMEN DE SEED")
        print("=" * 60)
        if dry_run:
            print(f"  Catalog entries (estimado) : {stats['catalog_entries']}")
            print(f"  Precios BRAME actuales     : {stats['brame_prices_current']}")
            print(f"  Precios BRAME pasados      : {stats['brame_prices_past']}")
        else:
            n_models   = db.query(VehicleModel).count()
            n_services = db.query(Service).count()
            n_catalog  = db.query(ServiceCatalog).count()
            n_brame_cur  = (db.query(SupplierPrice)
                .join(Supplier, SupplierPrice.supplier_id == Supplier.id)
                .filter(Supplier.name.ilike("brame"), SupplierPrice.is_current.is_(True))
                .count())
            n_brame_past = (db.query(SupplierPrice)
                .join(Supplier, SupplierPrice.supplier_id == Supplier.id)
                .filter(Supplier.name.ilike("brame"), SupplierPrice.is_current.is_(False))
                .count())
            print(f"  Modelos en BD              : {n_models}")
            print(f"  Servicios en BD            : {n_services}")
            print(f"  Entradas catálogo          : {n_catalog}")
            print(f"  Precios BRAME actuales     : {n_brame_cur}")
            print(f"  Precios BRAME pasados      : {n_brame_past}")
        print(f"  Errores                    : {total_errors}")
        print("=" * 60)
        print("[OK] Seed completado." if total_errors == 0 else "[WARN] Seed completado con errores.")

    except Exception as exc:
        db.rollback()
        print(f"\n[FATAL] {exc}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
